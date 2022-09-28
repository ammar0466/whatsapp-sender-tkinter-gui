import configparser
from multiprocessing.sharedctypes import Value
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo
import requests
import pyqrcodeng as pyqrcode
# from tqdm.auto import tqdm
import os
import time
import random
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import baileys_api 
import json
import qrcode

# import wwjs_api
from datetime import datetime   
import threading


# create the root window
root = tk.Tk()
root.title('Whatsapp Sender v1.07')
root.resizable(False, False)
root.geometry('400x550')

serverIp = "xxx.xxx.xxx.xxx"
serverPort = "7070"

def delaydulu():
    delay1 =textboxDelay.get("1.0", "end-1c")
    delay2 =textboxDelay2.get("1.0", "end-1c")
                
    time.sleep(random.randint(int(float(delay1)),int(float(delay2))))

def sendWa():
    senderNo = textboxSender.get("1.0", "end-1c")
    #trim senderNo
    senderNo = senderNo.strip()

    
    msgO = text_widget.get("1.0", "end-1c")
    
     

    # print (senderNo)
    #load workbook
    wb = load_workbook('wassap.xlsx')
    ws = wb.active
    maxPB = ws.max_row-1
    maxPBr = 1/maxPB*100
    maxC = ws.max_column
    if ws.cell(row=1, column=maxC).value == "status":
        maxCd = maxC-1 
    else:
        #add column name status
        ws.cell(row=1, column=maxC+1).value = "status"
        maxC=maxC+1
        maxCd = maxC-1

    # with tqdm(total=int(maxPB)) as progress_bar:
    #iterate over rows by header name
    # for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=maxCd):
         # reset msg
        msg = msgO
        # nameP = str(row[0].value)
        phoneNo = str(row[1].value)

        # find all value of column until max columns and replace to msg string
        # for cell in row:
        for cell in row[0:maxCd]:
        #limit iterate cell to maxC

            varCell = cell.value
            #check if varCell is datetime format
            if isinstance(varCell, datetime):
                varCell = varCell.strftime("%d/%m/%Y")


            # msg = msg.replace('<field'+str(cell)+'>', str(cell.value))
            msg = msg.replace('<field'+str(cell.column)+'>', str(varCell))
            msg = msg.replace('&', '%26')
        
        progress_bar['value']+=maxPBr
        #update tkinter label percent
        labelPercent.config(text=str(round(progress_bar['value'],2))+'%'+' '+'(' + str(row[0].row-1)+'/'+str(maxPB)+')')

        #print current iteration number
        # print(row[0].row)


        root.update_idletasks()
        
        #check if cell in column (status) is empty
        if ws.cell(row=row[0].row, column=maxC).value == None: 
            #progressbar
            
            # progress_bar.update(1)
            # print (message)
            

            

            #if empty, send message
            #guna whatsapp-web.js
            # response = wwjs_api.sendWhapi2(phoneNo, message)

            #guna whapi.io
            response = baileys_api.sendWhapi2(senderNo, phoneNo, msg)
            #update cell with current date
            ws.cell(row=row[0].row, column=maxC).value = response+str(datetime.now())
            # ws.cell(row=row[0].row, column=maxC+1).value = 
            # ws.cell(row=row[0].row, column=5).value = "response"
            
            #save workbook
            wb.save('wassap.xlsx')
            #update progress bar
            # add delay
            delaydulu()

        #else continue to next row
        else:
            # progress_bar.update(1)
            
            continue

        
        
        #update progress bar
        # progress['value'] = progress['value'] + 1
        # root.update_idletasks()


#create menu bar

menubar = tk.Menu(root)
filemenu = tk.Menu(menubar, tearoff=0)
filemenu.add_command(label="Open", command=lambda:openFile())
# filemenu.add_command(label="Save", command=lambda:saveFile())
filemenu.add_separator()
filemenu.add_command(label="Exit", command=root.quit)
menubar.add_cascade(label="File", menu=filemenu)
root.config(menu=menubar)

#add setting to menu bar
settingmenu = tk.Menu(menubar, tearoff=0)
settingmenu.add_command(label="Set Server", command=lambda:setting())
settingmenu.add_command(label="Add Sender", command=lambda:addSender())
menubar.add_cascade(label="Setting", menu=settingmenu)
root.config(menu=menubar)

#add help to menu bar
helpmenu = tk.Menu(menubar, tearoff=0)
helpmenu.add_command(label="About", command=lambda:about())
menubar.add_cascade(label="Help", menu=helpmenu)
root.config(menu=menubar)

#when clicked on Set Server open setting window
def setting():
    settingWindow = tk.Toplevel(root)
    settingWindow.title('Setting')
    settingWindow.resizable(False, False)
    settingWindow.geometry('400x200')
    #create label
    labelServer = tk.Label(settingWindow, text='Baileys-Api Server')
    labelServer.grid(row=0, column=0, padx=10, pady=10)
    #create textbox
    textboxServer = tk.Text(settingWindow, height=1, width=20)
    textboxServer.grid(row=0, column=1, padx=10, pady=10)
    #put placeholder text in textboxServer
    textboxServer.insert(tk.END, '207.909.234.234:7070')
    #create button
    buttonSave = tk.Button(settingWindow, text='Save', command=lambda:saveServer())
    buttonSave.grid(row=1, column=1, padx=10, pady=10)

    #save config textboxServer to file config.ini and close setting window when click on button save
    def saveServer():
        #save config textboxServer to file config.ini
        config = configparser.ConfigParser()
        config['server'] = {'apiServer': textboxServer.get("1.0", "end-1c")}
        with open('config.ini', 'w') as configfile:
            config.write(configfile)
        #close setting window
        settingWindow.destroy()

#when clicked on Add Sender open addSender window
def addSender():
    addSenderWindow = tk.Toplevel(root)
    addSenderWindow.title('Add Sender')
    addSenderWindow.resizable(False, False)
    #using grid layout


    # addSenderWindow.geometry('700x800')
    #create label
    labelSender = tk.Label(addSenderWindow, text='Sender Number')
    labelSender.grid(row=0, column=0, padx=10, pady=10)
    #create textbox
    textboxSender = tk.Text(addSenderWindow, height=1, width=16)
    textboxSender.grid(row=0, column=1, padx=10, pady=10)
    #put placeholder text in textboxSender
    textboxSender.insert(tk.END, '60123456789')
    #create button
    buttonSave = tk.Button(addSenderWindow, text='Close', command=lambda:closeSender())
    buttonSave.grid(row=2, column=1, padx=10, pady=10)

    buttonQR = tk.Button(addSenderWindow, text='Generate QR Code', command=lambda:generateQR())
    buttonQR.grid(row=1, column=1, padx=10, pady=10)

    #when click generate QR code, requests api to generate QR code
    def generateQR():
        global imge
        #get sender number from textboxSender
        senderNumber = textboxSender.get("1.0", "end-1c")

        # send senderID to baileys-api to get qr code
        urlQr = f"http://{serverIp}:{serverPort}/sessions/add"

        payload=f"id={senderNumber}&isLegacy=false"
        headers = {
            'Content-Type': 'application/x-www-form-urlencoded'
        }

        response = requests.request("POST", urlQr, headers=headers, data=payload)
        # save json response to python
        response = json.loads(response.text)
        # save json response to variable
        qr = response["data"]["qr"]
        #strip data:image/png;base64, from qr
        qr = qr.replace("data:image/png;base64,", "")
        # qr = "iVBORw0KGgoAAAANSUhEUgAAARQAAAEUCAYAAADqcMl5AAAAAklEQVR4AewaftIAABIpSURBVO3BQY7gRpIAQXei/v9l3z7GKQGCWS2NNszsD9Za64KHtda65GGttS55WGutSx7WWuuSh7XWuuRhrbUueVhrrUse1lrrkoe11rrkYa21LnlYa61LHtZa65KHtda65GGttS55WGutS374SOVvqjhROak4UZkqJpU3Kt5QmSomlaniC5Wp4kRlqnhDZaqYVKaKN1ROKr5QmSreUDmpmFT+poovHtZa65KHtda65GGttS754bKKm1S+qPhCZaqYVKaKm1SmikllqjhR+aLiDZWpYlKZKiaVk4qTiknli4oTlanipoqbVG56WGutSx7WWuuSh7XWuuSHX6byRsUbKlPFpDJVTConFScVk8pUMamcVEwqk8pUMalMFVPFpDKpvFFxk8obKm9UTCpTxaQyqUwVU8Wk8ptU3qj4TQ9rrXXJw1prXfKw1lqX/PAfo/JGxYnKVHFScVIxqZxUTCqTylRxojJVnKhMFScqJyonFZPKVPGFylQxqbyh8kbFpPJf8rDWWpc8rLXWJQ9rrXXJD/8xFZPKicobKlPFpHJScaJyUnGiclIxqbyhclIxqbyhcqIyVUwqU8VUcVIxqdykMlX8lzystdYlD2utdcnDWmtd8sMvq/ibVE5UTipOVCaVk4qTihOVE5WTiknlpOINlUllqjhRmSreUJkqJpWp4kRlqjhRmSomlanipop/k4e11rrkYa21LnlYa61LfrhM5Z9UMalMFZPKicpUMalMFZPKVDGpTBVfVEwqU8WkcqIyVZxUTCpTxRsqU8VNKlPFpDJV3KQyVZyo/Js9rLXWJQ9rrXXJw1prXWJ/8D9M5TdVfKHyRsWkMlWcqHxR8YbKVDGpnFR8oTJVnKi8UTGpvFHx/8nDWmtd8rDWWpc8rLXWJfYHH6hMFZPKTRVfqLxRMamcVLyhMlWcqLxRMan8popJ5TdVvKEyVUwqv6liUpkqJpWbKn7Tw1prXfKw1lqXPKy11iU/XKYyVUwqJxUnKlPFpHJS8ZtUpopJ5Q2VNypOKm5SOak4UTmpeEPlpopJ5Y2KSWWqmFRuqvibHtZa65KHtda65GGttS754S+reEPljYo3VKaKk4oTlZOKSeWNihOVqeJEZao4UZkqJpU3KiaVSWWqeKPiRGWqmFSmiptUpopJ5aTiC5Wp4ouHtda65GGttS55WGutS+wPPlCZKiaVqeILlTcqTlSmiknlpOJEZaqYVP5JFZPKb6qYVKaKN1TeqDhRmSomlaliUnmj4p+kMlV88bDWWpc8rLXWJQ9rrXXJD5epnKhMFZPKFxWTylTxRsWJyknFpDJVnKhMFScqU8Wk8kXFicpUMancVDGpTBWTylTxhcpUMalMFZPKVHGiMlWcqPxND2utdcnDWmtd8rDWWpfYH1yk8kbFpDJVTConFW+ofFHxhsobFf8mKicVX6hMFZPKFxX/JJWpYlI5qZhUTipOVKaKLx7WWuuSh7XWuuRhrbUusT/4i1RuqphUTireUDmpOFE5qfgnqXxRcaIyVUwqN1V8oXJScaLyRcWkclJxovJGxRcPa611ycNaa13ysNZal/zwkcobFZPKScWkclIxqUwqX1S8UXGiMlVMKr+p4kRlqphUpooTlZOKSWWqOFF5o+Kk4kTlpOKmijcqJpWp4qaHtda65GGttS55WGutS364rOJEZao4UXlDZaqYVN6omFSmikllqphUvqg4UZkqJpUTlanii4oTlUnlRGWqmComlaniDZWp4qRiUrlJZaqYVKaKqeI3Pay11iUPa611ycNaa11if/CBylTxhspU8YbKVDGpTBVvqEwVb6j8kyomlS8qvlB5o+INlTcq3lB5o+INlZOKE5WpYlKZKm56WGutSx7WWuuSh7XWuuSHjypOVKaKE5WTii9UpopJ5URlqjipOFE5qThRmSq+qJhUJpWpYlKZKqaKSeUNlanipOINlanipOILlaniRGWqOFGZKiaVqeKLh7XWuuRhrbUueVhrrUt++EhlqpgqTireUJkqTiomlTcqTlROKk4qTlSmihOVqWKqOFE5qfhC5aTipGJSmSomlaliUpkqTiomlaniRGWqmFSmijcqTlR+08Naa13ysNZalzystdYlP1ymMlV8ofKGym9SmSomlTdUpoqpYlI5qThROamYVCaVqWKquEnlpOKk4qRiUpkqJpUTlaliqphUpopJ5Q2VqWKq+E0Pa611ycNaa13ysNZal9gf/CKVqeJEZap4Q2WqmFSmir9JZaqYVE4qTlSmijdUpoq/SWWqmFSmihOVk4ovVE4q3lB5o+JE5aTipoe11rrkYa21LnlYa61LfvhIZaqYKt6omFSmii8q3lCZKiaVmyomlROVqeKLihOVqeILlaniDZWTikllUjmpOKmYVN5QOamYVG5SmSq+eFhrrUse1lrrkoe11rrkh48qJpWpYlJ5o+ImlaliUvmbVE4qJpUTlTcqJpWp4kTlpOJvqphUblJ5Q+ULlaniRGWqmFR+08Naa13ysNZalzystdYlP3ykMlVMKlPFGypvVLyhclIxqXxRMalMFTdVvFExqUwVX6icqEwVJxWTyknFP6niROVEZaqYKiaVqWJSuelhrbUueVhrrUse1lrrkh/+YSonFW+onFScqJxUfKEyVUwqJxWTylQxqZxUTCpfqEwVU8WkMlWcVJxUTCqTylQxqUwVX1T8JpWp4kTlNz2stdYlD2utdcnDWmtdYn/wgcobFScqJxWTyhcVN6mcVJyoTBWTyk0VJypvVJyoTBWTylRxovJGxRcqU8WJylQxqUwVk8pJxYnKGxVfPKy11iUPa611ycNaa13yw0cVX6hMFScqb1RMKicqU8Wk8jepTBWTylQxqUwVJyonFZPKicpUcVLxRsWkcqIyVUwqU8UbKm9UTConFScqU8Wk8pse1lrrkoe11rrkYa21LrE/uEhlqjhR+Zsq3lCZKk5U3qiYVN6omFSmihOVmyreUJkqvlCZKiaVqWJSOamYVE4qTlRuqphU3qj44mGttS55WGutSx7WWuuSHz5SeUPlpOINlZOKSeWkYqo4UZkqJpWbKiaVN1ROKt5QeUPlb1J5o+JEZaqYVL6oeENlUpkqJpWp4qaHtda65GGttS55WGutS364rGJS+UJlqvibVL6omFROKk5U3lD5QmWqOFF5o2JS+aJiUnlDZao4UTlRmSreUJkqTipOKiaVqeKLh7XWuuRhrbUueVhrrUt++MsqJpWTijcqJpU3VN6oOFF5Q2WqmComlTcqJpWTijcq3lCZKr5QmSomlUllqviiYlKZVN6ouEllqrjpYa21LnlYa61LHtZa6xL7g79I5W+qmFSmihOVk4oTlZOKN1ROKt5QuaniROWkYlKZKt5QmSreUPmiYlL5N6m46WGttS55WGutSx7WWuuSHz5SOamYKr5QOamYVE5U3qj4N1P5ouJE5TepTBWTyhcqJxVTxRsqJxVfqEwVJypTxW96WGutSx7WWuuSh7XWuuSHyypOVKaKSeWkYlJ5o2JSmSomlZsqvqiYVE4qJpWp4o2KSeWNihOVSeWkYlI5qfhNFTepTBWTylRxonJS8cXDWmtd8rDWWpc8rLXWJfYHH6icVJyoTBVvqEwVv0llqphU3qi4SWWqmFSmijdUpoovVKaKSeWk4kRlqvhCZaqYVG6q+F/ysNZalzystdYlD2utdYn9wUUqU8UXKm9UTCpTxaRyUnGTyknFpHJS8YXKGxWTylQxqZxUfKHyN1VMKicVk8pU8YbKVPGFylTxxcNaa13ysNZalzystdYlP3yk8psqJpWpYlI5UTmpmFSmihOVqWKqmFQmlTdU3qh4o+KLijdUpopJ5YuKSWWqOFGZKiaVk4pJZar4QmWqmFR+08Naa13ysNZalzystdYlP3xUMalMKlPFGypTxaTymyq+UHmj4g2VN1TeUJkqpoo3VKaKNyomlaniROVEZar4m1T+poqbHtZa65KHtda65GGttS754SOVk4pJZaqYVKaKmypOVN6oOKk4Ufmi4kTlpOJEZVKZKr5QmSomlaniDZUvVE5UflPFicpUcVIxqUwVXzystdYlD2utdcnDWmtd8sM/TGWqmFSmipOKSWVSmSpOKv5JKlPFpPKFyhsVJyo3VZxUvFHxhsobFZPKScUbKicqU8Xf9LDWWpc8rLXWJQ9rrXXJD79M5aTipOINlZOKN1ROKk5UblKZKr5QeUPljYo3VKaKSeWk4qaKE5VJZar4QuWk4t/kYa21LnlYa61LHtZa6xL7g4tUpopJ5YuKN1SmikllqnhDZaqYVE4q/iaVk4pJ5Y2KSWWqmFSmiknlpGJSmSpOVKaKL1ROKk5U3qiYVKaKSeWk4ouHtda65GGttS55WGutS374l6u4SeUNlanii4pJZaqYVE4qTlSmikllUjmpuKliUjmpmFTeUJkqTlSmiptUbqo4qfhND2utdcnDWmtd8rDWWpfYH/yDVKaKSeWNihOVqeINlaniRGWqeENlqjhRmSreUPknVUwqJxVvqEwVk8oXFScqU8WJyhsVk8obFV88rLXWJQ9rrXXJw1prXWJ/8ItUpoqbVN6omFR+U8WkMlVMKlPFpPJGxaRyUnGiclIxqdxUcaIyVXyhMlWcqEwVk8pUMam8UTGpTBV/08Naa13ysNZalzystdYl9gcfqJxUnKhMFZPKScWJyknFpDJVnKjcVPGGyknFFypfVNykclIxqZxUTCpTxaTyRsWk8kXFicpUMamcVHzxsNZalzystdYlD2utdckPH1VMKpPKScVJxYnKVPGGylQxqUwVJxVvqJyofKEyVUwqb1R8oTJVTCpTxUnFScWJylQxqUwVX1T8lz2stdYlD2utdcnDWmtdYn/wF6n8TRUnKl9UfKEyVZyoTBWTylRxojJVTConFZPKVHGiMlVMKlPFicobFZPKTRWTyhsVk8pUMamcVPymh7XWuuRhrbUueVhrrUvsDy5SOak4UZkqvlA5qXhD5Y2KE5WpYlKZKiaVqeINlS8q3lB5o+ILlaniN6lMFW+ovFHxhcpU8cXDWmtd8rDWWpc8rLXWJfYHv0jlv6RiUvmiYlK5qWJSmSreUJkqJpU3Kk5UTiomlaliUjmpmFSmihOVqeJEZaqYVN6omFTeqPjiYa21LnlYa61LHtZa65IfPlJ5o2JSmSp+k8pJxaRyU8UXFScqk8pU8YbKVPFGxaRyojJVvFExqXxRMamcVEwqU8WJylQxqbxRMalMFTc9rLXWJQ9rrXXJw1prXfLDX6YyVZyoTBVfVJyoTBW/SWWquKliUvlC5SaVE5Wp4kTljYpJ5aaKSeUNld+kMlV88bDWWpc8rLXWJQ9rrXWJ/cH/MJW/qeJEZaqYVKaKN1TeqDhRmSreUJkq3lCZKk5Uvqj4QuWNiknlpOINlanin/Sw1lqXPKy11iUPa611yQ8fqfxNFVPFpPJGxaRyojJVTBWTylQxqUwVk8pJxYnKVPGGylRxojJVnFRMKm9UTCpTxRsqU8UbFScVk8qJylRxovJFxRcPa611ycNaa13ysNZal/xwWcVNKicqJxVvVEwqJypTxT9JZar4ouImlZOKSWWq+EJlqnij4g2VLyreqDhRmSpuelhrrUse1lrrkoe11rrkh1+m8kbFFxWTylTxRsWkMlVMKlPFScWkMlXcpHKicpPKVDGpvKEyVZyoTBUnFScqJxUnFZPKpPKbKn7Tw1prXfKw1lqXPKy11iU//Meo/E0qU8Wk8kbFpHJTxYnKGxVvqJyoTBU3qUwVX1ScVEwqU8WJyknFFypTxRcPa611ycNaa13ysNZal/zwH1MxqUwqU8Wk8m9ScZPKScWkMlWcqEwVU8WkMlWcqEwVU8WkMlW8oXKi8kbFicr/soe11rrkYa21LnlYa61LfvhlFb+p4o2Kk4pJZao4UXmjYlK5qeKkYlL5TSpTxaRyUnGicqJyUjFVTCpvVPxNKlPF3/Sw1lqXPKy11iUPa611yQ+XqfxNKm+oTBWTylQxqfymiknljYqTipOKSWVSmSreqJhUporfVHGi8kXFicobFZPKScWk8jc9rLXWJQ9rrXXJw1prXWJ/sNZaFzystdYlD2utdcnDWmtd8rDWWpc8rLXWJQ9rrXXJw1prXfKw1lqXPKy11iUPa611ycNaa13ysNZalzystdYlD2utdcnDWmtd8n+PP318ZlVrRwAAAABJRU5ErkJggg=="

        status = response["success"] 
        message = response["message"] #MESSAGE SUCCESSFULLY SENT TO INDEX.HTML

        #print response to console
        # print (response)
        #print qr last 10 character
        print (qr[-10:])

        imge = tk.PhotoImage(data=qr)
        #output qr code to tkinter label
        qrCodeWindow = tk.Toplevel(addSenderWindow)
        labelQRt = tk.Label(qrCodeWindow, text='Scan QR Code')
        
        labelQr = tk.Label(qrCodeWindow, image=imge)
        labelQRt.pack(pady=10)
        labelQr.pack(pady=10)

        #clear imge
        

        
        #save qr code to file
        # im.write('qr.png')
        


        
        

        
    
       
        

        # # return qr to index.html
        # # return {"status": status, "message": message}
        # return {"status": status, "qrCode": qr,  "message": message}


        # #convert response to json
        # response = response.json()
        # #get qr code from response
        # qrCode = response['qrCode']
        # #convert qr code to image
        # qrCode = pyqrcode.create(qr)

        # #popup qrcode image
        # qrCodeWindow = tk.Toplevel(addSenderWindow)
        # qrCodeWindow.title('QR Code')
        # qrCodeWindow.resizable(False, False)
        # qrCodeWindow.geometry('300x300')
        # #create label
        # labelQR = tk.Label(qrCodeWindow, text='Scan QR Code')
        # labelQR.grid(row=0, column=0, padx=10, pady=10)
        # #create image
        # qrCodeImage = qrCode.png('qrCode.png', scale=8)
        
        # labelQRCode = tk.Label(qrCodeWindow, image=qrCodeImage)
        # labelQRCode.grid(row=1, column=0, padx=10, pady=10)


        # #span qr code image to addSenderWindow
        # #span qr code image to addSenderWindow
        
        # qrCodeImage = tk.Label(addSenderWindow, image=qrCode)
        # qrCodeImage.grid(row=3, column=1, padx=10, pady=10)
        # # qrCodeImage.pack()
        # qrCodeImage.image = qrCode




        

        

    #save config textboxSender to file config.ini and close addSender window when click on button save
    def closeSender():
        
        addSenderWindow.destroy()
    








#create label
labelSender = tk.Label(root, text="Sender Number 601xxxxx (Must)")
#create textbox
textboxSender = tk.Text(root, height=1, width=14)


#create label
labelDelay = tk.Label(root, text="Delay Random from - to in seconds (Must)")
labelDelay2 = tk.Label(root, text="-")

#create textbox
textboxDelay = tk.Text(root, height=1, width=10)
textboxDelay2 = tk.Text(root, height=1, width=10)

check_button = ttk.Button(
    root,
    text='Send',
    command=threading.Thread(target=sendWa).start
)

progress_bar = ttk.Progressbar(root, style='text.Horizontal.TProgressbar', length=300, mode='determinate')

labelPercent = ttk.Label(root, text="0%")

#create label
labelMsg = tk.Label(root, text="Write message, eg: <field1>.., <field2> is column Receiver Number") 

#create label
labelFile = tk.Label(root, text="Default file wassap.xlsx on same directory") 
labelFormat = tk.Label(root, text="format contoh: field1 = Name, field2 = Phone No., last column = status") 
# Create the text widget
text_widget = tk.Text(root, height=15, width=40)
 

 


labelSender.pack()
textboxSender.pack()
labelDelay.pack()
textboxDelay.pack()
labelDelay2.pack()
textboxDelay2.pack()
labelMsg.pack()
 
text_widget.pack()
check_button.pack(pady=10)
progress_bar.pack(pady=5)
labelPercent.pack()
labelFile.pack()
labelFormat.pack()




# run the application
root.mainloop()
